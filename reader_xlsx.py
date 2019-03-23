import xlrd
import openpyxl
import re
import Order
import Machine
from datetime import datetime

class products:
    _id = []
    equipment_class = []


def ConvertToUnix(date):
    down = date.split("-")
    uma = []
    for i in down:
        uma.append(int(i))
    return 31556926*(uma[0]-1970)+uma[1]*2629743+uma[2]*86400

"""class order:
    _id = []
    product_id = []
    amount = []
    deadline = []"""

"""class equipment:
    _id = []
    _class = []
    available_hours = []
    speed_per_hour = []"""
PATH = '/home/picknick/Wallpapers/test/'
 # i
=======
def read_order():
    # reading products
    prod = products()
    rb = openpyxl.load_workbook(filename = PATH + 'product.xlsx')
    sheet = rb.active
    for i in range(sheet.max_row - 1):
        buf = sheet.cell(row = i + 2, column = 1).value
        prod._id.append(buf)
        buf = sheet.cell(row = i + 2, column = 2).value
        buf = re.findall(r"[\w\s-]+" ,buf)
        buf_list = []
        for st in buf:
            if st != ' ':
                buf_list.append(st)
        prod.equipment_class.append(buf_list)
        """for st in buf:
            prod.equipment_class.append(st)"""

    #reading order
    prod_1 = []
    rb = openpyxl.load_workbook(filename = PATH + 'order.xlsx')
    sheet = rb.active
    for i in range(sheet.max_row - 1):
        #buf = sheet.cell(row = i + 1, column = 1).value
        #prod._id.append(buf)
        t = Order.Order(0,0,[0])
        prod_1.append(t)
        buf = sheet.cell(row = i + 2, column = 2).value
        for i1 in range(len(products._id)):
            if buf == products._id[i1]:
                prod_1[i].nodeArray =  prod.equipment_class[i1]
        buf = sheet.cell(row = i + 2, column = 3).value
        prod_1[i].amount = int(buf)
        buf = sheet.cell(row = i + 2, column = 4).value
        prod_1[i].deadline = ConvertToUnix(buf)
    return prod_1

def read_machine():
    #reading equipment
    prod = []
    t = Machine.Machine('','', 0, 0.0)
    rb = openpyxl.load_workbook(filename = PATH + 'equipment.xlsx')
    sheet = rb.active
    for i in range(sheet.max_row):
        prod.append(t)
        buf = sheet.cell(row = i + 1, column = 1).value
        prod[i]._id = buf
        buf = sheet.cell(row = i + 1, column = 2).value
        prod[i]._type = buf
        buf = sheet.cell(row = i + 1, column = 4).value
        prod[i].speed = buf
    return prod
    