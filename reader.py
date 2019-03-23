import xlrd
import openpyxl
import re

class products:
    _id = []
    equipment_class = []

class order:
    _id = []
    product_id = []
    amount = []
    deadline = []

class equipment:
    _id = []
    _class = []
    available_hours = []
    speed_per_hour = []


# reading products
prod = products()
rb = xlrd.open_workbook(r'C:\Users\USer\Desktop\product_1.xls', formatting_info=True)
sheet = rb.sheet_by_index(0)
for i in range(sheet.nrows):
    buf = sheet.row_values(i)[0]
    prod._id.append(buf)
    buf = sheet.row_values(i)[1]
    buf = re.findall(r"[\w ]+", buf)
    buf_list = []
    for st in buf:
        buf_list.append(st)
    prod.equipment_class.append(buf_list)
    """for st in buf:
        prod.equipment_class.append(st)"""
print(prod._id[1])
for st in prod.equipment_class[1]:
    print(st)
#
# # reading products
# prod = products()
# PATH = input("Enter the path to product xlsx table")
# rb = xlrd.open_workbook(PATH, formatting_info = True)
# sheet = rb.sheet_by_index(0)
# for i in range(sheet.nrows):
#     buf = sheet.row_values(i)[0]
#     prod._id.append(buf)
#     buf = sheet.row_values(i)[1]
#     buf = re.findall(r"[\w ]+" ,buf)
#     buf_list = []
#     for st in buf:
#         buf_list.append(st)
#     prod.equipment_class.append(buf_list)
#     """for st in buf:
#         prod.equipment_class.append(st)"""
# print(prod._id[1])
# for st in prod.equipment_class[1]:
#     print(st)
#
# #reading order
# prod = order()
# PATH = input("Enter the path to order xlsx table")
# rb = openpyxl.load_workbook(filename = PATH)
# sheet = rb.active
# for i in range(sheet.max_row):
#     buf = sheet.cell(row = i + 1, column = 1).value
#     prod._id.append(buf)
#     buf = sheet.cell(row = i + 1, column = 2).value
#     prod.product_id.append(buf)
#     buf = sheet.cell(row = i + 1, column = 3).value
#     prod.amount.append(buf)
#     buf = sheet.cell(row = i + 1, column = 4).value
#     prod.deadline.append(buf)
# print(prod._id[1])
#
# #reading equipment
# prod = equipment()
# PATH = input("Enter the path to equip xlsx table")
# rb = openpyxl.load_workbook(filename = PATH)
# sheet = rb.active
# for i in range(sheet.max_row):
#     buf = sheet.cell(row = i + 1, column = 1).value
#     prod._id.append(buf)
#     buf = sheet.cell(row = i + 1, column = 2).value
#     prod._class.append(buf)
#     buf = sheet.cell(row = i + 1, column = 4).value
#     prod.speed_per_hour.append(buf)
# print(prod._id[1])
